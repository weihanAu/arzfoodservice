import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
import pandas as pd
import os

def load_file():
    global file_path
    file_path = filedialog.askopenfilename(
        title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        label.config(text=f"Loaded: {os.path.basename(file_path)}")

def is_valid(value):
    return value is not None and value != '' and bool(value)

def duplicate_rows_based_on_quantity(file_path):
    # Load the workbook and select the first sheet
    wb = load_workbook(file_path)
    ws = wb.active  # Adjust this if you need a specific sheet

    # Find the last row
    last_row = ws.max_row

    # Loop through rows starting from the last row and working upwards
    for i in range(last_row, 2, -1):  # Starts at row 3
        qty = ws.cell(row=i, column=7).value  # Adjust to the correct column for Quantity
        unit = ws.cell(row=i, column=8).value  # Get the box unit info
        isWeight = ws.cell(row=i, column=9).value  # Check if this item is a weight item
        print(isWeight)
        if is_valid(isWeight):
            print(isWeight)
            continue

        if not isinstance(qty, (int, float)) or not isinstance(unit, (int, float)):
            raise ValueError(f"Error: Row {i} has invalid data types. "
                             f"Quantity: {qty}, Unit: {unit} (Both should be integers).")

        if qty == unit:
            ws.cell(row=i, column=7).value = "1 box"

        while qty > unit:
            ws.cell(row=i, column=7).value = "1 box"
            ws.insert_rows(i + 1)
            for j in range(1, ws.max_column + 1):
                ws.cell(row=i + 1, column=j).value = ws.cell(row=i, column=j).value

            ws.cell(row=i + 1, column=7).value = "1 box"
            ws.cell(row=i + 1, column=8).value = unit
            qty -= unit

        if 0 < qty < unit:
            ws.cell(row=i, column=7).value = f"{qty} unit" if qty == 1 else f"{qty} units"

    base_name, ext = os.path.splitext(file_path)  # Split file name and extension
    new_file_name = f"{base_name}_labels{ext}"   # Add '_labels' to the name
    wb.save(new_file_name)

    # Load the Excel file
    df = pd.read_excel(new_file_name, engine='openpyxl', skiprows=1)
    # Save the DataFrame to a CSV file
    csv_file = f"{base_name}_labels.csv"  # Change this to your desired output file name
    df.to_csv(csv_file, index=False)
    
    os.remove(f"{base_name}_labels.xlsx")


def generate_labels():
    if not file_path:
        messagebox.showwarning("Warning", "Please load an Excel file first.")
        return

    try:
        duplicate_rows_based_on_quantity(file_path)
        messagebox.showinfo("Success", f"Labels file created: {os.path.basename(file_path).replace('.xlsx', '_labels.csv')}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Initialize the application
app = tk.Tk()
app.title("Label Generator - ARZ foodservice")

file_path = ""

app.geometry("400x400")

# Create a frame for the UI with padding
frame = tk.Frame(app, padx=20, pady=20)
frame.pack(padx=10, pady=10)

# Load File Button
load_button = tk.Button(app, text="Load Excel File", command=load_file)
load_button.pack(pady=10)

# Label to display loaded file
label = tk.Label(app, text="")
label.pack(pady=10)

# Generate Labels Button
generate_button = tk.Button(app, text="Generate Labels", command=generate_labels)
generate_button.pack(pady=10)

# Text below the Generate Labels button
info_label = tk.Label(app, text=
    "Please don't modify the CSV file downloaded from SAGE. Coz it has the Default structure as below: \n"
    "\n"
    "row 1 is the CSV file title meta.\n"
    "\n"
    "row 2 is the column description.\n"
    "\n"
    "Column G is the quantity that the customer ordered.\n"
    "\n"
    "Column H is the number of items in one box.\n"
    "\n"
    "Column I has the value 'TRUE' is for weighted items.\n"
    "\n",
      wraplength=300)
info_label.pack(pady=5)


app.mainloop()
