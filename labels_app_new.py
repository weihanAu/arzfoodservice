import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import pandas as pd
import os
import re

big_array =[
    "EGGS - 700g  {59's}",
    "EGGS - 600g  {55's} {SUNEGGS}",
    "EGGS FREE RANGE - 700g  **** TRAY BOX ****"
    ]

small_array=[
    'SHANKLISH CHEESE {Appox 240g} Rw',
    'GLOVES VYNAL LARGE x 100 (10)',
    'PAPER TOWELS (16)',
    'GLOVES VYNAL LARGE x 100 (10)',
    "CONTAINERS RECTANGLE **** 500ml **** 50pk (10)",
    "CONTAINERS RECTANGLE **** 650ml **** 50pk (10)",
    "CONTAINERS RECTANGLE **** 750ml **** 50pk (10)",
    "LIDS FOR **** RECTANGLE **** CONTAINERS 50pk (10)",
    "HEAVY DUTY CONTAINERS **** 1000ml ****  50pk (10)",
    "HEAVY DUTY CONTAINERS **** 750ml ****  50pk (10)",
    "LIDS FOR RECTANGLE **** HEAVY DUTY **** CONTAINERS 50pk (10)",
    "CONTAINER PLASTIC ROUND **** 280ml ****  50pk (10)",
    "CONTAINER PLASTIC ROUND **** 440ml **** x 50 (10)",
    "LIDS FOR **** ROUND **** CONTAINERS 50pk (10)",
    "CONTAINER PLASTIC ROUND WITH LID 70ml x 50 (20)",
    "CONTAINER PLASTIC ROUND 70ml x 50 (20)",
    "LIDS x 100 **** 70ml CONTAINER PLASTIC ROUND **** (10)"
    ]

def load_file():
    global file_path
    file_path = filedialog.askopenfilename(
        title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        label.config(text=f"Loaded: {os.path.basename(file_path)}")

def is_valid(value):
    return value is not None and value != '' and bool(value)

def duplicate_rows_based_on_quantity(file_path):
    # Load the workbook and select the first sheet
    wb = load_workbook(file_path)
    ws = wb.active  # Adjust this if you need a specific sheet

    # Find the last row
    last_row = ws.max_row

    # Loop through rows starting from the last row and working upwards
    for i in range(last_row, 2, -1):  # Starts at row 3
        
        #remove the row if its a frozen item,
        location_code = ws.cell(row=i, column=5).value
        if location_code.startswith("F"):
            ws.delete_rows(i)
            continue

        canPatch = False
        qty_description = ws.cell(row=i, column=10).value
        number_in_parentheses = re.search(r'\((\d+)\)', qty_description)
        if number_in_parentheses:
             number = number_in_parentheses.group(1)
             number = int(number)
             if number:
              canPatch = True
              unit = number
       
        issmallitem = ws.cell(row=i, column=11).value  # Adjust to the correct column for Quantity
        qty = ws.cell(row=i, column=12).value  # Adjust to the correct column for Quantity

        #check unit weight in description
        item_weight = re.search(r'(\d+\.\d+|\d+)kg', qty_description)

        if item_weight:
            if float(item_weight.group(1))<=2:
             print(float(item_weight.group(1)))
             issmallitem = True

        if  re.search(r'(\d+)g', qty_description):
            issmallitem = True
      
        # check if it has strings like '6 x 100 pack'
        if  re.search(r'x\s*\d+', qty_description, re.IGNORECASE):  # Case-insensitive search
            issmallitem = False
        # additions check for some special items
        if qty_description in big_array:
            issmallitem = False
        if qty_description in small_array:
            issmallitem = True

        # if canPatch is False, means don't modify the current row
        if canPatch == False:
            unit = 1 #default unit
            if is_valid(issmallitem):
               if qty == 1:
                ws.cell(row=i, column=12).value=f"{qty}unit" 
               if qty > 1:
                ws.cell(row=i, column=12).value=f"{qty}units" 
               continue
       
        if not isinstance(qty, (int, float)) or not isinstance(unit, (int, float)):
            raise ValueError(f"Error: Row {i} has invalid data types. "
                             f"Quantity: {qty}, Unit: {unit} (Both should be integers).")
        # if unit per box == 1
        if unit == 1:
            if qty ==1:
                ws.cell(row=i, column=12).value = "1 unit"
            while qty > 1:
                ws.cell(row=i, column=12).value = "1 unit"
                ws.insert_rows(i + 1)
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=i + 1, column=j).value = ws.cell(row=i, column=j).value

                ws.cell(row=i + 1, column=12).value = "1 unit"
                ws.cell(row=i + 1, column=13).value = unit
                qty -= 1
            # then stop the codes running    
            continue
        # if unit > 1
        if qty == unit:
            ws.cell(row=i, column=12).value = "1 box"

        while qty > unit:
            ws.cell(row=i, column=12).value = "1 box"
            ws.insert_rows(i + 1)
            for j in range(1, ws.max_column + 1):
                ws.cell(row=i + 1, column=j).value = ws.cell(row=i, column=j).value

            ws.cell(row=i + 1, column=12).value = "1 box"
            ws.cell(row=i + 1, column=13).value = unit
            qty -= unit

        if qty == 1:
               ws.cell(row=i, column=12).value = "1 unit"

        if 1 < qty and qty < unit:
             if is_valid(issmallitem):
                ws.insert_rows(i + 1)  # 插入新行，避免覆盖
                for j in range(1, ws.max_column + 1):  # 复制当前行内容
                        ws.cell(row=i + 1, column=j).value = ws.cell(row=i, column=j).value
                ws.cell(row=i+1, column=12).value = f"{qty} unit" if qty == 1 else f"{qty} units"

             if qty > 1 and not is_valid(issmallitem):
                  while qty > 0 and qty < unit:
                    ws.insert_rows(i + 1)  # 插入新行，避免覆盖
                    for j in range(1, ws.max_column + 1):  # 复制当前行内容
                        ws.cell(row=i + 1, column=j).value = ws.cell(row=i, column=j).value
                    # 更新新行的数据
                    ws.cell(row=i + 1, column=12).value = "1 unit"
                    ws.cell(row=i + 1, column=13).value = unit
                    qty = qty - 1  # 递减 qty
             ws.delete_rows(i)

    base_name, ext = os.path.splitext(file_path)  # Split file name and extension
    new_file_name = f"{base_name}_labels{ext}"   # Add '_labels' to the name
    wb.save(new_file_name)

    # Load the Excel file
    df = pd.read_excel(new_file_name, engine='openpyxl', skiprows=1)
    # reverse the csv file
    df_reversed = df.iloc[::-1]
    #sort the rows
    df_reversed = df_reversed.sort_values(by=df_reversed.columns[5], kind='mergesort')
    # Save the DataFrame to a CSV file
    csv_file = f"{base_name}_labels.csv"  # Change this to your desired output file name
    df_reversed.to_csv(csv_file, index=False)
    
    os.remove(f"{base_name}_labels.xlsx")

def generate_labels():
    if not file_path:
        messagebox.showwarning("Warning", "Please load an Excel file first.")
        return

    try:
        duplicate_rows_based_on_quantity(file_path)
        messagebox.showinfo("Success", f"Labels file created: {os.path.basename(file_path).replace('.xlsx', '_labels.csv')}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Initialize the application
app = tk.Tk()
app.title("Label Generator - ARZ foodservice")

file_path = ""

app.geometry("400x400")

# Create a frame for the UI with padding
frame = tk.Frame(app, padx=20, pady=20)
frame.pack(padx=11, pady=11)

# Load File Button
load_button = tk.Button(app, text="Load Excel File", command=load_file)
load_button.pack(pady=11)

# Label to display loaded file
label = tk.Label(app, text="")
label.pack(pady=11)

# Generate Labels Button
generate_button = tk.Button(app, text="Generate Labels", command=generate_labels)
generate_button.pack(pady=11)

# Text below the Generate Labels button
info_label = tk.Label(app, text=
    "Please don't modify the CSV file downloaded from SAGE. Coz it has the Default structure as below: \n"
    "\n"
    "row 1 is the CSV file title meta.\n"
    "\n"
    "row 2 is the column description.\n"
    "\n"
    "Column G is the quantity that the customer ordered.\n"
    "\n"
    "Column H is the number of items in one box.\n"
    "\n"
    "Column I has the value 'TRUE' is for weighted items.\n"
    "\n",
      wraplength=300)
info_label.pack(pady=5)


app.mainloop()
